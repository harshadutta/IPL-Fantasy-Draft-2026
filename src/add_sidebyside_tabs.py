"""Add side-by-side XI comparison tabs to the Excel workbook."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import glob

wb = load_workbook('output/IPL_Fantasy_Draft_2026.xlsx')

TEAM_ORDER = ['CSK', 'MI', 'RCB', 'KKR', 'DC', 'PBKS', 'RR', 'SRH', 'GT', 'LSG']
THIN_BORDER = Border(bottom=Side(style='thin', color='D0D0D0'))
SUBHEADER_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

TEAM_FILLS = {
    'CSK': PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid'),
    'MI':  PatternFill(start_color='004BA0', end_color='004BA0', fill_type='solid'),
    'RCB': PatternFill(start_color='C62828', end_color='C62828', fill_type='solid'),
    'KKR': PatternFill(start_color='3D0066', end_color='3D0066', fill_type='solid'),
    'DC':  PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid'),
    'PBKS': PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid'),
    'RR':  PatternFill(start_color='E91E90', end_color='E91E90', fill_type='solid'),
    'SRH': PatternFill(start_color='FF6F00', end_color='FF6F00', fill_type='solid'),
    'GT':  PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid'),
    'LSG': PatternFill(start_color='00897B', end_color='00897B', fill_type='solid'),
}
TEAM_FONT_CSK = Font(bold=True, size=12, color='000000')  # dark text on yellow
TEAM_FONT_DEFAULT = Font(bold=True, size=12, color='FFFFFF')

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RED_FILL = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

# Load data
inet = pd.read_csv('data/processed/internet_predicted_xis.csv')
xi_files = glob.glob('data/processed/playing_xi_*.csv')
research = pd.concat([pd.read_csv(f) for f in xi_files], ignore_index=True)

# Remove old versions if they exist
for name in ['XIs Side-by-Side (Internet)', 'XIs Side-by-Side (Research)']:
    if name in wb.sheetnames:
        del wb[name]

# ── Tab: Internet XIs side-by-side ──
ws1 = wb.create_sheet(title='XIs Side-by-Side (Internet)', index=2)
CPT = 3  # columns per team: #, Player, Role
GAP = 1

for i, team in enumerate(TEAM_ORDER):
    sc = i * (CPT + GAP) + 1  # start column

    # Row 1: Team name header
    cell = ws1.cell(row=1, column=sc, value=team)
    cell.font = TEAM_FONT_CSK if team == 'CSK' else TEAM_FONT_DEFAULT
    cell.fill = TEAM_FILLS.get(team, SUBHEADER_FILL)
    cell.alignment = Alignment(horizontal='center')
    ws1.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc + CPT - 1)

    # Row 1 match info
    team_df = inet[inet['team'] == team].sort_values('batting_pos')
    match_info = team_df.iloc[0].get('match_info', '') if len(team_df) > 0 else ''
    cell2 = ws1.cell(row=2, column=sc, value=match_info)
    cell2.font = Font(italic=True, size=9, color='555555')
    ws1.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=sc + CPT - 1)

    # Row 3: Sub-headers
    for j, h in enumerate(['#', 'Player', 'Role']):
        c = ws1.cell(row=3, column=sc + j, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal='center')

    # Rows 4-14: Players
    for row_idx, (_, pl) in enumerate(team_df.head(11).iterrows()):
        r = row_idx + 4
        ws1.cell(row=r, column=sc, value=pl['batting_pos']).alignment = Alignment(horizontal='center')
        name = pl['player']
        if str(pl.get('nationality', '')).strip() == 'Overseas':
            name += ' *'
        ws1.cell(row=r, column=sc + 1, value=name)
        ws1.cell(row=r, column=sc + 2, value=pl['role']).alignment = Alignment(horizontal='center')
        for j in range(CPT):
            ws1.cell(row=r, column=sc + j).border = THIN_BORDER

    # Source row
    source = team_df.iloc[0].get('source', '') if len(team_df) > 0 else ''
    ws1.cell(row=16, column=sc, value=source).font = Font(italic=True, size=8, color='888888')

    # Column widths
    ws1.column_dimensions[get_column_letter(sc)].width = 4
    ws1.column_dimensions[get_column_letter(sc + 1)].width = 22
    ws1.column_dimensions[get_column_letter(sc + 2)].width = 14
    if sc + CPT <= 256:
        ws1.column_dimensions[get_column_letter(sc + CPT)].width = 2

ws1.freeze_panes = 'A4'

# ── Tab: Researched XIs side-by-side ──
ws2 = wb.create_sheet(title='XIs Side-by-Side (Research)', index=3)
CPT2 = 5  # #, Player, Role, Start%, Overs
GAP2 = 1

for i, team in enumerate(TEAM_ORDER):
    sc = i * (CPT2 + GAP2) + 1

    # Row 1: Team name
    cell = ws2.cell(row=1, column=sc, value=team)
    cell.font = TEAM_FONT_CSK if team == 'CSK' else TEAM_FONT_DEFAULT
    cell.fill = TEAM_FILLS.get(team, SUBHEADER_FILL)
    cell.alignment = Alignment(horizontal='center')
    ws2.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc + CPT2 - 1)

    # Row 2: Sub-headers
    for j, h in enumerate(['#', 'Player', 'Role', 'P(Start)', 'Overs']):
        c = ws2.cell(row=2, column=sc + j, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal='center')

    # Filter to likely players
    team_df = research[
        (research['team'] == team) & (research['prob_starting'] >= 40)
    ].sort_values('expected_batting_pos').head(15)

    for row_idx, (_, pl) in enumerate(team_df.iterrows()):
        r = row_idx + 3
        ws2.cell(row=r, column=sc, value=int(pl['expected_batting_pos'])).alignment = Alignment(horizontal='center')

        name = pl['player']
        nat = str(pl.get('nationality', ''))
        if nat in ['Overseas', 'OS']:
            name += ' *'
        ws2.cell(row=r, column=sc + 1, value=name)
        ws2.cell(row=r, column=sc + 2, value=pl['role']).alignment = Alignment(horizontal='center')

        prob = int(pl['prob_starting'])
        ws2.cell(row=r, column=sc + 3, value=f"{prob}%").alignment = Alignment(horizontal='center')
        ws2.cell(row=r, column=sc + 4, value=pl['expected_overs_bowled']).alignment = Alignment(horizontal='center')

        # Color by probability
        if prob >= 90:
            fill = GREEN_FILL
        elif prob >= 70:
            fill = YELLOW_FILL
        else:
            fill = RED_FILL

        for j in range(CPT2):
            ws2.cell(row=r, column=sc + j).fill = fill
            ws2.cell(row=r, column=sc + j).border = THIN_BORDER

    # Column widths
    ws2.column_dimensions[get_column_letter(sc)].width = 4
    ws2.column_dimensions[get_column_letter(sc + 1)].width = 22
    ws2.column_dimensions[get_column_letter(sc + 2)].width = 14
    ws2.column_dimensions[get_column_letter(sc + 3)].width = 9
    ws2.column_dimensions[get_column_letter(sc + 4)].width = 6
    if sc + CPT2 <= 256:
        ws2.column_dimensions[get_column_letter(sc + CPT2)].width = 2

ws2.freeze_panes = 'A3'

try:
    wb.save('output/IPL_Fantasy_Draft_2026.xlsx')
except PermissionError:
    wb.save('output/IPL_Fantasy_Draft_2026_v2.xlsx')
    print('(saved as v2 because original was open)')
print('Done! Added side-by-side tabs.')
print('Tabs:', [ws.title for ws in wb.worksheets])
