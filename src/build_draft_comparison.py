"""
Build Excel workbook comparing all draft teams with stats and commentary.
"""

import pandas as pd
import numpy as np
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Styling
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
THIN_BORDER = Border(bottom=Side(style='thin', color='D0D0D0'))
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RED_FILL = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
USER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
BOLD = Font(bold=True)


def load_all_data():
    """Load concept board and 2025 actuals."""
    board = pd.read_csv('output/concept_draft_board.csv')
    actuals = pd.read_csv('data/processed/player_2025_actuals.csv')
    return board, actuals


def find_player_stats(player_name, ipl_team, board, actuals):
    """Find player stats, trying multiple name formats."""
    # Try exact match in concept board
    match = board[board['player'] == player_name]
    if len(match) == 0:
        # Try partial match
        for _, row in board.iterrows():
            if player_name.lower() in row['player'].lower() or row['player'].lower() in player_name.lower():
                match = board[board['player'] == row['player']]
                break

    # Try actuals with Cricsheet names
    act_match = actuals[actuals['player'] == player_name]
    if len(act_match) == 0:
        # Try last name match
        last = player_name.split()[-1].lower()
        first_init = player_name.split()[0][0].upper() if player_name.split() else ''
        for _, row in actuals.iterrows():
            p = row['player']
            p_last = p.split()[-1].lower()
            p_first = p.split()[0][0].upper() if p.split() else ''
            if p_last == last and p_first == first_init:
                act_match = actuals[actuals['player'] == p]
                break
        # Special cases
        name_map = {
            'B Sai Sudharsan': 'B Sai Sudharsan',
            'KL Rahul': 'KL Rahul',
            'Vaibhav Suryavanshi': 'V Suryavanshi',
            'MS Dhoni': 'MS Dhoni',
            'Quinton de Kock': 'Q de Kock',
            'Tim David': 'TH David',
            'Phil Salt': 'PD Salt',
            'Nitish Kumar Reddy': 'Nithish Kumar Reddy',
        }
        if player_name in name_map:
            mapped = name_map[player_name]
            m2 = actuals[actuals['player'] == mapped]
            if len(m2) > 0:
                act_match = m2

    result = {
        'avg_pts_2025': None,
        'total_pts_2025': None,
        'matches_2025': 0,
        'blended_pts': None,
        'pred_season': None,
        'concept_total': None,
        'role': '',
        'prob_starting': None,
    }

    if len(match) > 0:
        r = match.iloc[0]
        result['blended_pts'] = r.get('blended_pts_per_match')
        result['pred_season'] = r.get('pred_season_total')
        result['concept_total'] = r.get('concept_total')
        result['role'] = r.get('role', '')
        result['prob_starting'] = r.get('prob_starting')
        result['avg_pts_2025'] = r.get('avg_pts_2025')
        result['total_pts_2025'] = r.get('total_pts_2025')
        result['matches_2025'] = r.get('matches_2025', 0)

    if len(act_match) > 0:
        a = act_match.iloc[0]
        if result['avg_pts_2025'] is None:
            result['avg_pts_2025'] = a.get('avg_pts_2025')
        if result['total_pts_2025'] is None:
            result['total_pts_2025'] = a.get('total_pts_2025')
        if result['matches_2025'] == 0:
            result['matches_2025'] = int(a.get('matches_2025', 0))

    return result


def auto_width(ws, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def main():
    # Load data
    with open('data/processed/draft_teams.json', 'r') as f:
        teams = json.load(f)

    board, actuals = load_all_data()

    wb = Workbook()
    wb.remove(wb.active)

    # ── Tab per team ──
    team_summaries = []

    for team in teams:
        team_name = team['name']
        is_user = team.get('is_user', False)
        ws = wb.create_sheet(title=team_name[:31])  # Excel tab name limit

        # Headers
        headers = ['#', 'Player', 'IPL Team', 'Role Tag', 'Role',
                   'Avg Pts 2025', 'Total Pts 2025', 'Matches 2025',
                   'Concept Pts', 'Blended Pts/Match', 'Pred Season',
                   'P(Start)%']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        total_pred = 0
        total_2025 = 0
        players_with_data = 0
        roles = []

        for idx, player in enumerate(team['players']):
            row = idx + 2
            stats = find_player_stats(player['player'], player['ipl_team'], board, actuals)

            ws.cell(row=row, column=1, value=idx + 1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=player['player']).font = BOLD if player.get('role_tag') else Font()
            ws.cell(row=row, column=3, value=player['ipl_team']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=player.get('role_tag', ''))

            role = stats['role'] or ''
            ws.cell(row=row, column=5, value=role)
            roles.append(role)

            # 2025 actuals
            avg25 = stats['avg_pts_2025']
            total25 = stats['total_pts_2025']
            matches25 = stats['matches_2025']
            ws.cell(row=row, column=6, value=round(avg25, 1) if avg25 else '').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=7, value=int(total25) if total25 and not np.isnan(total25) else '').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8, value=matches25 if matches25 else '').alignment = Alignment(horizontal='center')

            # Concept / blended
            ws.cell(row=row, column=9, value=round(stats['concept_total'], 1) if stats['concept_total'] else '').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=10, value=round(stats['blended_pts'], 1) if stats['blended_pts'] else '').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=11, value=round(stats['pred_season'], 0) if stats['pred_season'] else '').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=12, value=round(stats['prob_starting'], 0) if stats['prob_starting'] else '').alignment = Alignment(horizontal='center')

            # Color by avg pts
            if avg25 and avg25 >= 70:
                fill = GREEN_FILL
            elif avg25 and avg25 >= 45:
                fill = YELLOW_FILL
            elif avg25 and avg25 > 0:
                fill = RED_FILL
            else:
                fill = PatternFill()  # no fill for unknowns

            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
                if avg25:
                    ws.cell(row=row, column=c).fill = fill

            # Accumulate
            if stats['pred_season']:
                total_pred += stats['pred_season']
            if total25 and not (isinstance(total25, float) and np.isnan(total25)):
                total_2025 += total25
                players_with_data += 1

        # Summary row
        sum_row = len(team['players']) + 3
        ws.cell(row=sum_row, column=1, value='TOTAL').font = BOLD
        ws.cell(row=sum_row, column=7, value=int(total_2025)).font = BOLD
        ws.cell(row=sum_row, column=11, value=int(total_pred)).font = BOLD

        avg_avg = total_2025 / players_with_data if players_with_data > 0 else 0
        ws.cell(row=sum_row + 1, column=1, value='Avg per player (2025 actual)').font = Font(italic=True)
        ws.cell(row=sum_row + 1, column=6, value=round(avg_avg, 1)).font = BOLD

        auto_width(ws)
        ws.freeze_panes = 'A2'

        # Collect summary
        team_summaries.append({
            'team': team_name,
            'is_user': is_user,
            'total_pred_season': int(total_pred),
            'total_2025_pts': int(total_2025),
            'players_with_2025': players_with_data,
            'avg_2025_per_player': round(avg_avg, 1),
            'roles': roles,
        })

    # ── Summary comparison tab ──
    ws_sum = wb.create_sheet(title='Team Comparison', index=0)

    headers = ['Rank', 'Team', 'Pred Season Total', 'Total 2025 Pts',
               'Players w/ 2025 Data', 'Avg 2025 Pts/Player']
    for col, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Sort by predicted season total
    team_summaries.sort(key=lambda x: x['total_pred_season'], reverse=True)

    for idx, ts in enumerate(team_summaries):
        row = idx + 2
        ws_sum.cell(row=row, column=1, value=idx + 1).alignment = Alignment(horizontal='center')
        cell_name = ws_sum.cell(row=row, column=2, value=ts['team'])
        ws_sum.cell(row=row, column=3, value=ts['total_pred_season']).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=row, column=4, value=ts['total_2025_pts']).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=row, column=5, value=ts['players_with_2025']).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=row, column=6, value=ts['avg_2025_per_player']).alignment = Alignment(horizontal='center')

        if ts['is_user']:
            cell_name.font = Font(bold=True, color='0066CC')
            for c in range(1, len(headers) + 1):
                ws_sum.cell(row=row, column=c).fill = USER_FILL

        for c in range(1, len(headers) + 1):
            ws_sum.cell(row=row, column=c).border = THIN_BORDER

    auto_width(ws_sum)
    ws_sum.freeze_panes = 'A2'

    # Save
    out_path = 'output/Draft_Teams_Comparison.xlsx'
    wb.save(out_path)
    print(f'Saved: {out_path}')
    print(f'Teams: {len(teams)}')
    print()

    # Print summary
    print('=== TEAM RANKING BY PREDICTED SEASON TOTAL ===')
    for idx, ts in enumerate(team_summaries):
        marker = ' <<<< YOU' if ts['is_user'] else ''
        print(f"  {idx+1}. {ts['team']:20s}  Pred: {ts['total_pred_season']:>6,}  2025 actual: {ts['total_2025_pts']:>6,}  Avg/player: {ts['avg_2025_per_player']:>5.1f}{marker}")


if __name__ == '__main__':
    main()
