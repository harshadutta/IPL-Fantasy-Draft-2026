"""
Build the final Excel draft workbook with multiple tabs.
"""

import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ── Styling ──
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT_WHITE = Font(bold=True, size=11, color='FFFFFF')
TIER1_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
TIER2_FILL = PatternFill(start_color='D6F0FF', end_color='D6F0FF', fill_type='solid')
TIER3_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
CLIFF_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='D0D0D0')
)

TIER_FILLS = {
    1: PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
    2: PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    3: PatternFill(start_color='D6F0FF', end_color='D6F0FF', fill_type='solid'),
    4: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    5: PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),
    6: PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
    7: PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
}


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def style_data_rows(ws, df, nrows, ncols, tier_col=None, cliff_col=None):
    for row_idx in range(2, nrows + 2):
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center') if col_idx <= 3 else Alignment(horizontal='left')

        # Tier coloring
        if tier_col is not None:
            tier_val = ws.cell(row=row_idx, column=tier_col).value
            if tier_val in TIER_FILLS:
                for col_idx in range(1, ncols + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = TIER_FILLS[tier_val]

        # Cliff highlighting
        if cliff_col is not None:
            cliff_val = ws.cell(row=row_idx, column=cliff_col).value
            if cliff_val == True or cliff_val == 'True' or cliff_val == 'TRUE':
                for col_idx in range(1, ncols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = CLIFF_FILL
                    cell.font = Font(bold=True)


def auto_width(ws, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def add_df_to_sheet(wb, sheet_name, df, tier_col_name=None, cliff_col_name=None):
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append([v if not (isinstance(v, float) and np.isnan(v)) else '' for v in r])

    ncols = len(df.columns)
    nrows = len(df)
    style_header(ws, ncols)

    tier_col = None
    cliff_col = None
    if tier_col_name and tier_col_name in df.columns:
        tier_col = list(df.columns).index(tier_col_name) + 1
    if cliff_col_name and cliff_col_name in df.columns:
        cliff_col = list(df.columns).index(cliff_col_name) + 1

    style_data_rows(ws, df, nrows, ncols, tier_col, cliff_col)
    auto_width(ws)
    ws.freeze_panes = 'A2'
    return ws


def build_concept_patterns():
    """Analyze which structural concept blocks scored the most in 2025."""
    fp = pd.read_csv('data/processed/player_match_fantasy_points.csv')

    # Load batting and bowling ball-level data if available
    bat_concepts = None
    bowl_concepts = None
    if os.path.exists('output/batting_concept_ranking.csv'):
        bat_concepts = pd.read_csv('output/batting_concept_ranking.csv')
    if os.path.exists('output/bowling_concept_ranking.csv'):
        bowl_concepts = pd.read_csv('output/bowling_concept_ranking.csv')

    # Build match-level concept patterns
    # Group by role archetype and show who scored most
    def role_arch(row):
        bp = row['batting_pos']
        ov = row['overs_bowled']
        if bp == 0 and ov == 0:
            return 'Sub/Unused'
        elif ov >= 3 and 1 <= bp <= 5:
            return 'Batting AR (top 5, 3+ ov)'
        elif ov >= 3 and (bp > 5 or bp == 0):
            return 'Pure Bowler (3+ ov)'
        elif 1 <= ov < 3 and 1 <= bp <= 5:
            return 'Bat-leaning AR (top 5, 1-2 ov)'
        elif 1 <= ov < 3 and bp > 5:
            return 'Bowl-leaning AR (lower, 1-2 ov)'
        elif 1 <= bp <= 2 and ov < 1:
            return 'Pure Opener (1-2)'
        elif bp == 3 and ov < 1:
            return 'One-down (3)'
        elif 4 <= bp <= 5 and ov < 1:
            return 'Middle-order Bat (4-5)'
        elif 6 <= bp <= 7 and ov < 1:
            return 'Lower-middle Bat (6-7)'
        elif bp >= 8 and ov < 3:
            return 'Tail/Part-timer'
        else:
            return 'Other'

    fp['concept_slot'] = fp.apply(role_arch, axis=1)
    fp = fp[fp['concept_slot'] != 'Sub/Unused']

    # Summary by concept slot
    concept_summary = fp.groupby('concept_slot').agg(
        player_matches=('total_fantasy_pts', 'count'),
        unique_players=('player', 'nunique'),
        avg_fantasy_pts=('total_fantasy_pts', 'mean'),
        median_fantasy_pts=('total_fantasy_pts', 'median'),
        std_fantasy_pts=('total_fantasy_pts', 'std'),
        floor_p10=('total_fantasy_pts', lambda x: np.percentile(x, 10)),
        ceiling_p90=('total_fantasy_pts', lambda x: np.percentile(x, 90)),
        avg_batting_pts=('batting_pts', 'mean'),
        avg_bowling_pts=('bowling_pts', 'mean'),
        avg_fielding_pts=('fielding_pts', 'mean'),
        avg_balls_faced=('balls_faced', 'mean'),
        avg_overs_bowled=('overs_bowled', 'mean'),
    ).round(1).sort_values('avg_fantasy_pts', ascending=False)

    concept_summary['batting_share'] = (concept_summary['avg_batting_pts'] / concept_summary['avg_fantasy_pts'] * 100).round(0)
    concept_summary['bowling_share'] = (concept_summary['avg_bowling_pts'] / concept_summary['avg_fantasy_pts'] * 100).round(0)

    # Top players per concept slot
    top_per_concept = []
    for slot in concept_summary.index:
        slot_df = fp[fp['concept_slot'] == slot]
        player_avgs = slot_df.groupby('player').agg(
            matches=('total_fantasy_pts', 'count'),
            avg_pts=('total_fantasy_pts', 'mean'),
            total_pts=('total_fantasy_pts', 'sum'),
        ).query('matches >= 5').sort_values('avg_pts', ascending=False)

        for i, (player, row) in enumerate(player_avgs.head(5).iterrows()):
            top_per_concept.append({
                'concept_slot': slot,
                'rank_in_slot': i + 1,
                'player': player,
                'matches': int(row['matches']),
                'avg_pts': round(row['avg_pts'], 1),
                'total_pts': int(row['total_pts']),
            })

    top_df = pd.DataFrame(top_per_concept)

    return concept_summary, top_df, bat_concepts, bowl_concepts


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Tab 1: Draft Board ──
    print("Building Tab 1: Draft Board...")
    board = pd.read_csv('output/concept_draft_board.csv')

    # Ensure total_pts_2025 is included
    board_cols = [
        'rank', 'tier', 'cliff', 'player', 'team', 'role', 'nationality',
        'expected_batting_pos', 'expected_overs_bowled',
        'avg_pts_2025', 'total_pts_2025', 'matches_2025',
        'median_pts_2025', 'floor_p10_2025', 'ceiling_p90_2025',
        'avg_batting_pts_2025', 'avg_bowling_pts_2025',
        'concept_total', 'blended_pts_per_match', 'pred_season_total',
        'expected_games', 'prob_starting', 'prob_all_14',
        'is_returning', 'notes',
    ]
    board_cols = [c for c in board_cols if c in board.columns]
    board_out = board[board_cols].copy()
    # Filter to only players with prob_starting >= 50 for cleaner board
    board_main = board_out[board_out['prob_starting'] >= 50].copy()
    board_main['rank'] = range(1, len(board_main) + 1)

    add_df_to_sheet(wb, 'Draft Board', board_main, tier_col_name='tier', cliff_col_name='cliff')

    # ── Tab 2: Full Player List (all 202) ──
    print("Building Tab 2: All Players...")
    add_df_to_sheet(wb, 'All Players', board_out, tier_col_name='tier')

    # ── Tab 3: Concept Patterns ──
    print("Building Tab 3: Concept Patterns...")
    concept_summary, top_per_concept, bat_concepts, bowl_concepts = build_concept_patterns()
    concept_summary_reset = concept_summary.reset_index().rename(columns={'concept_slot': 'Concept Slot'})
    add_df_to_sheet(wb, 'Concept Patterns', concept_summary_reset)

    # ── Tab 4: Top Players per Concept ──
    print("Building Tab 4: Top per Concept...")
    add_df_to_sheet(wb, 'Top per Concept', top_per_concept)

    # ── Tab 5: Batting Concepts ──
    if bat_concepts is not None:
        print("Building Tab 5: Batting Concepts...")
        bat_cols = ['bat_rank', 'concept', 'n', 'avg_bat_pts', 'median_bat_pts', 'avg_balls_faced', 'avg_sr']
        bat_cols = [c for c in bat_cols if c in bat_concepts.columns]
        add_df_to_sheet(wb, 'Batting Concepts', bat_concepts[bat_cols])

    # ── Tab 6: Bowling Concepts ──
    if bowl_concepts is not None:
        print("Building Tab 6: Bowling Concepts...")
        bowl_cols = ['bowl_rank', 'concept', 'n', 'avg_bowl_pts', 'median_bowl_pts', 'avg_overs', 'avg_wickets', 'avg_dots', 'avg_er']
        bowl_cols = [c for c in bowl_cols if c in bowl_concepts.columns]
        add_df_to_sheet(wb, 'Bowling Concepts', bowl_concepts[bowl_cols])

    # ── Tab 7: Venue Analysis ──
    print("Building Tab 7: Venue Analysis...")
    if os.path.exists('output/descriptives/points_by_venue.csv'):
        venue = pd.read_csv('output/descriptives/points_by_venue.csv')
        add_df_to_sheet(wb, 'Venue Analysis', venue)

    # ── Tab 8: 2025 Actuals Leaderboard ──
    print("Building Tab 8: 2025 Leaderboard...")
    actuals = pd.read_csv('data/processed/player_2025_actuals.csv')
    actuals = actuals.sort_values('total_pts_2025', ascending=False)
    actuals['rank'] = range(1, len(actuals) + 1)
    act_cols = ['rank', 'player', 'team_2025', 'matches_2025', 'total_pts_2025', 'avg_pts_2025',
                'median_pts_2025', 'std_pts_2025', 'floor_p10_2025', 'ceiling_p90_2025',
                'avg_batting_pts', 'avg_bowling_pts', 'avg_fielding_pts',
                'typical_bat_pos_2025', 'avg_overs_bowled_2025']
    act_cols = [c for c in act_cols if c in actuals.columns]
    add_df_to_sheet(wb, '2025 Leaderboard', actuals[act_cols].head(100))

    # ── Tab 9: Predicted XIs (internet sources) ──
    print("Building Tab 9: Internet Predicted XIs...")
    if os.path.exists('data/processed/internet_predicted_xis.csv'):
        preds = pd.read_csv('data/processed/internet_predicted_xis.csv')
        add_df_to_sheet(wb, 'Internet Predicted XIs', preds)
    else:
        print("  (internet predictions not yet available, will add later)")

    # ── Tab 10: Researched XIs with probabilities ──
    print("Building Tab 10: Researched Starting XIs...")
    import glob
    xi_files = glob.glob('data/processed/playing_xi_*.csv')
    if xi_files:
        xi_dfs = [pd.read_csv(f) for f in xi_files]
        all_xis = pd.concat(xi_dfs, ignore_index=True)
        # Filter to likely starters
        starters = all_xis[all_xis['prob_starting'] >= 60].sort_values(['team', 'expected_batting_pos'])
        add_df_to_sheet(wb, 'Researched XIs', starters)

    # Save
    out_path = 'output/IPL_Fantasy_Draft_2026.xlsx'
    wb.save(out_path)
    print(f"\nSaved Excel workbook: {out_path}")
    print(f"Tabs: {[ws.title for ws in wb.worksheets]}")


if __name__ == '__main__':
    main()
